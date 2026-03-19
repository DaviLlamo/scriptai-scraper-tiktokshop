from openpyxl import load_workbook
from ia import gerar_resposta  # importa a função da IA

def analisar_excel(caminho):
    wb = load_workbook(caminho)
    ws = wb.active

    resultados = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        nome, link = row

        if not nome:
            continue

        prompt = f"""
Você é especialista em vídeos virais do TikTok Shop.

Crie 2 roteiros para:

Título: {nome}

Inclua:
- Gancho
- Desenvolvimento
- CTA
- Linguagem viral
"""
        texto = gerar_resposta(prompt)

        resultados.append({
            "nome": nome,
            "link": link,
            "roteiros": texto
        })

    return resultados